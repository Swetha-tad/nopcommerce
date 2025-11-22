import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def getColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file,sheetname,row_num,col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.cell(row=row_num,column=col_num).value)

def writeData(file,sheetname,row_num,col_num,value):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.cell(row=row_num,column=col_num).value)

def fillGreenColor(file,sheetname,row_num,col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenFill=PatternFill(start_color='00FF00',end_color='00FF00',fill_type = 'solid')
    sheet.cell(row_num,col_num).fill=greenFill
    workbook.save(file)

def fillRedColor(file,sheetname,row_num,col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redFill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type = 'solid')
    sheet.cell(row_num,col_num).fill = redFill
    workbook.save(file)

